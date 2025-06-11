import json
import tempfile
import uuid
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def upload_to_vercel_blob(file_path, blob_name):
    token = os.environ.get("BLOB_READ_WRITE_TOKEN")
    if not token:
        raise ValueError("Missing BLOB_READ_WRITE_TOKEN")

    file_name = f"meal_plans/{blob_name}.xlsx"

    # Step 1: Request an upload URL from Vercel
    res = requests.post(
        "https://api.vercel.com/v2/blob/upload-url",
        headers={"Authorization": f"Bearer {token}"},
        json={"filename": file_name}
    )
    res.raise_for_status()
    upload_info = res.json()

    # Step 2: Upload the file to that URL
    with open(file_path, "rb") as f:
        put_res = requests.put(upload_info["url"], data=f)
        put_res.raise_for_status()

    # Step 3: Return the public download URL
    return upload_info["url"].split("?")[0]

def handler(request):
    # Handle CORS preflight
    if request.method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
            },
            'body': ''
        }

    # Only allow POST requests
    if request.method != 'POST':
        return {
            'statusCode': 405,
            'headers': {'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Method not allowed'})
        }

    try:
        # Parse request body
        if hasattr(request, 'get_json'):
            data = request.get_json()
        else:
            body = getattr(request, 'body', None) or getattr(request, 'data', None)
            if body:
                if isinstance(body, bytes):
                    body = body.decode('utf-8')
                data = json.loads(body)
            else:
                return {
                    'statusCode': 400,
                    'headers': {'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({'error': 'No data provided'})
                }

        calorie_target = data["calorie_target"]
        protein_target = data["protein_target"]
        days = data["days"]

        # Create workbook
        wb = Workbook()

        for day, meals in days.items():
            ws = wb.create_sheet(title=day)

            # Summary
            ws.append(["Metric", "Target", "Actual", "Difference"])
            ws.append(["Calories (kcal)", calorie_target, "", ""])
            ws.append(["Protein (grams)", protein_target, "", ""])
            ws.append([])

            total_day_cal = 0
            total_day_protein = 0

            for meal in meals:
                ws.append([meal["meal_name"]])
                ws.append(["Meal", "Ingredient", "Quantity", "Unit", "Protein (g)", "Calories"])

                meal_cal = 0
                meal_protein = 0

                for ing in meal["ingredients"]:
                    ws.append([
                        meal["meal_name"],
                        ing["name"],
                        ing["quantity"],
                        ing["unit"],
                        ing["protein"],
                        ing["calories"]
                    ])
                    meal_cal += ing["calories"]
                    meal_protein += ing["protein"]

                ws.append(["TOTAL", "", "", "", meal_protein, meal_cal])
                ws.append([])

                total_day_cal += meal_cal
                total_day_protein += meal_protein

            ws.append(["Daily Totals", "", "", "", total_day_protein, total_day_cal])

            # Update summary cells
            ws["C2"] = total_day_cal
            ws["D2"] = total_day_cal - calorie_target
            ws["C3"] = total_day_protein
            ws["D3"] = total_day_protein - protein_target

            # Styling
            for row in ws.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    cell.font = Font(bold=True)

            for row in ws.iter_rows(min_row=1, max_col=4, max_row=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Remove default sheet if present
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Save file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"meal_plan_{uuid.uuid4()}.xlsx")
        wb.save(file_path)

        # Upload to blob
        blob_name = f"meal_plan_{uuid.uuid4()}"
        download_url = upload_to_vercel_blob(file_path, blob_name)

        # Cleanup
        os.remove(file_path)

        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*',
            },
            'body': json.dumps({"downloadUrl": download_url})
        }

    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': str(e)})
        }

