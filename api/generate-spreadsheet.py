from http.server import BaseHTTPRequestHandler
import json
import tempfile
import uuid
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Get content length and read body
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                self.send_error(400, "No data provided")
                return
            
            # Read and parse JSON data
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            calorie_target = data["calorie_target"]
            protein_target = data["protein_target"]
            days = data["days"]
            
            # Create workbook
            wb = Workbook()
            
            for day, meals in days.items():
                ws = wb.create_sheet(title=day)
                
                # Summary row at top
                ws.append(["Metric", "Target", "Actual", "Difference"])
                ws.append(["Calories (kcal)", calorie_target, "", ""])
                ws.append(["Protein (grams)", protein_target, "", ""])
                ws.append([])
                
                total_day_cal = 0
                total_day_protein = 0
                
                for meal in meals:
                    ws.append([meal["meal_name"]])
                    ws.append(["Meal", "Ingredient", "Quantity", "Unit", "Protein (g)", "Calories"])
                    
                    meal_cal = 0
                    meal_protein = 0
                    
                    for ing in meal["ingredients"]:
                        ws.append([
                            meal["meal_name"],
                            ing["name"],
                            ing["quantity"],
                            ing["unit"],
                            ing["protein"],
                            ing["calories"]
                        ])
                        meal_cal += ing["calories"]
                        meal_protein += ing["protein"]
                    
                    ws.append(["TOTAL", "", "", "", meal_protein, meal_cal])
                    ws.append([])
                    
                    total_day_cal += meal_cal
                    total_day_protein += meal_protein
                
                ws.append(["Daily Totals", "", "", "", total_day_protein, total_day_cal])
                
                # Update summary
                ws["C2"] = total_day_cal
                ws["D2"] = total_day_cal - calorie_target
                ws["C3"] = total_day_protein
                ws["D3"] = total_day_protein - protein_target
                
                # Optional formatting
                for row in ws.iter_rows(min_row=1, max_row=3):
                    for cell in row:
                        cell.font = Font(bold=True)
                
                for row in ws.iter_rows(min_row=1, max_col=4, max_row=3):
                    for cell in row:
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            
            # Save to temporary file
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, f"meal_plan_{uuid.uuid4()}.xlsx")
            wb.save(file_path)
            
            # Send response
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="meal_plan.xlsx"')
            self.end_headers()
            
            # Send file content
            with open(file_path, 'rb') as f:
                self.wfile.write(f.read())
            
            # Clean up
            os.remove(file_path)
            
        except Exception as e:
            self.send_error(500, f"Server error: {str(e)}")
    
    def do_GET(self):
        self.send_error(405, "Method not allowed")
